# -*- coding: utf-8 -*-
"""
Created on Mon Dec 17 08:56:26 2018

@author: Ronny Sentongo

Created by	: Ronny Sentongo, Enfogroup AB
Contact*  	: Robert.Granfors@enfo.se, Ronny.Sentongo@enfogroup.com
Purpose*  	: Robert Granfors (Kruxet är att deras CFO tar denna rapport och vill exportera den till Excel för att få in den i ett större rapportpaket han gör.
                              Problemet med denna rapport, är att när man exporterar den till Excel så måste jag eller deras CFO manuellt lägga rätt mycket tid 
                              för att den ska passa in i mallen som också är bifogad.)
Costumer 	: BRA flyg
Output	  	: 
loggin      :
--Ladda hem VPN-klienten cisco connect
--Anslut till: extvpn.flygbra.se
--User: rogr
--Pass: Vinter2019

----PWRBI loggin

Server:         ALSSQLP012.mcn.malmoaviation.se
Användarnamn:   DWH-reader
Lösenord:       BM@vat1000 




Notes



This is a Python-script, therefore some downloads/installation will be needed
1. You need to have PowerBI downloaded from Microsoft Store
2. Download and install Python 3 https://www.python.org/downloads/
3. Install Pip, if not included, make sure that you computer recognise Pip
    controlpanel -system-avanced envorialment-envorialment variables- path edit (then add new path) these path below

  
4. Now install the modules your using (se import): The is needed for PWRBI

"""
#C:\Users\Admin\AppData\Local\Programs\Python\Python37-32
#C:\Users\Admin\AppData\Local\Programs\Python\Python37-32\Scripts

#..import all modules needed...................................................

import pyodbc
import pandas as pd
import sqlite3
import time
import os
from openpyxl import load_workbook
from datetime import datetime






Today=time.strftime("%Y-%m-%d")
start_time = time.time()

df=pd.DataFrame({'Datex':['0000-01-01']})
df['Date']=datetime.strptime(time.strftime("%Y-%m-%d"), '%Y-%m-%d')
df['WeekNo']= df.Date.dt.strftime('%W')
df['period_start'] = ((df.Date - pd.DateOffset(months=1)).dt.to_period("m").dt.start_time).dt.strftime('%Y%m01')
df['FirstDateOfMonth']= (df.Date.dt.to_period("m").dt.start_time).dt.strftime('%Y%m01')
df['EndPeriod'] = (df.Date + pd.DateOffset(months=2)).dt.to_period("m").dt.start_time
df['LastDateOfMonth_N']= (df.EndPeriod.dt.to_period("m").dt.end_time).dt.strftime('%Y%m%d')
df['period_startDate'] = ((df.Date - pd.DateOffset(months=2)).dt.to_period("m").dt.start_time).dt.strftime("%Y-%m-01")
df['period_EndDate'] = ((df.Date + pd.DateOffset(months=12)).dt.to_period("m").dt.start_time).dt.strftime("%Y-%m-01")
df['Period1'] = ((df.Date - pd.DateOffset(months=0)).dt.to_period("m").dt.start_time).dt.strftime('%Y%m')
df['Period2'] = ((df.Date + pd.DateOffset(months=1)).dt.to_period("m").dt.start_time).dt.strftime('%Y%m')
df['Period3'] = ((df.Date + pd.DateOffset(months=2)).dt.to_period("m").dt.start_time).dt.strftime('%Y%m')
df['x'] = ((df.Date - pd.DateOffset(months=0)).dt.to_period("m").dt.start_time)
df['Period1_excel'] = ((df.x - pd.DateOffset(months=0)).dt.to_period("m").dt.start_time).dt.strftime('%B %Y')
df['Period2_excel'] = ((df.x + pd.DateOffset(months=1)).dt.to_period("m").dt.start_time).dt.strftime('%B %Y')
df['Period3_excel'] = ((df.x + pd.DateOffset(months=2)).dt.to_period("m").dt.start_time).dt.strftime('%B %Y')







period_start="[]".join(list(df.period_start))
period_start_hTable="[]".join(list(df.FirstDateOfMonth))
period_end_hTable="[]".join(list(df.LastDateOfMonth_N))
period_startDate="[]".join(list(df.period_startDate)) 
period_EndDate="[]".join(list(df.period_EndDate))
period_to_excel=list(df.Period1)+list(df.Period2)+list(df.Period3)
week="[]".join(list(df.WeekNo))
Period1_excel="[]".join(list(df.Period1_excel))
Period2_excel="[]".join(list(df.Period2_excel))
Period3_excel="[]".join(list(df.Period3_excel))







                         
chuck_sum=0
chunksize=1000*100
print('Program begins' )

#..........difine the costumer and the path for your report....................
costumer='BRA'
path=r'C:\Users\F2531197\Enfo Oyj\Robert Granfors - BRA flyg-Ronny'

lista='powerbi database outfiles'

folders = [item for item in lista.title().split(' ') if  item not in  ['']] #exclude empty


#create the folders for the rapport

for num, folder_name in enumerate(folders, start=1):
    newpath = path+'\\'+folder_name
    if not os.path.exists(newpath):
        os.makedirs(newpath)

#.................this is the connection to the costumer.......................
       
#connection in
server = 'ALSSQLP012.mcn.malmoaviation.se'
database = 'bradw_dm'
username = 'DWH-reader'
password = 'BM@vat1000'

#SQL Server 
SQLServer = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};' 
                              'Server=ALSSQLP012.mcn.malmoaviation.se;' 
                              'Database=bradw_dm;' 
                              'uid=DWH-reader;' 
                              'pwd=BM@vat1000')

SQLServer_cursor = SQLServer.cursor()


#create a database

SQLITE3 = sqlite3.connect(path+'\\Database\\'+costumer+'.db')      

SQLITE3_cursor = SQLITE3.cursor()



'''............................................................................
#..............................................................................
#....read the tables exactly how they were read in PWRBI....................'''

'''.............................read dim tables............................'''

lista='vdCitypair vdExchangeRate'

tables = [item for item in lista.split(' ') if  item not in  ['']]

for table in tables:
    sql_query= """  select * from pbiFB.%s"""%(table)
    pd.io.sql.execute('DROP table IF EXISTS '+table, SQLITE3)
    for df in pd.read_sql_query(sql_query, SQLServer,chunksize=chunksize): #get data from SQLITE
        df.to_sql(name=table,con=SQLITE3,index=False,if_exists='append')


'''.............................read fact tables............................'''


table='vfFlights'
sql_query= """

SELECT
     *
FROM
    pbiFB.vfFlights vf
where DepartureDate_ID>=%s and AreaRegionCode>999
    
    """%(period_start)
pd.io.sql.execute('DROP TABLE IF EXISTS '+ table, SQLITE3)
for df in pd.read_sql_query(sql_query, SQLServer,chunksize=chunksize):
    df.to_sql(name=table,con=SQLITE3,index=False,if_exists='append')


table='fFlights'
sql_query= """

SELECT
      vcp.CitypairBKEY,
      vcp.CitypairName,
      vcp.CitypairSortOrder,
      vcp.CitypairOriginCountryCode,
      vcp.CitypairDestinationCountryCode,
      vf.DepartureDate_ID,
      vf.SeatCountTotal
FROM
    vfFlights vf
    join vdCitypair vcp ON vcp.Citypair_ID = vf.Citypair_ID
    
    """

pd.io.sql.execute('DROP TABLE IF EXISTS '+ table, SQLITE3)
for df in pd.read_sql_query(sql_query, SQLITE3,chunksize=chunksize):
    df['Date'] =(df.DepartureDate_ID.astype(str).str[0:4]+'-'+
    df.DepartureDate_ID.astype(str).str[4:6]+'-'+
    df.DepartureDate_ID.astype(str).str[6:8]).values.astype("datetime64[s]")
    df['DomesticOrInternational']='International'
    df.loc[(df.CitypairOriginCountryCode == 'SE') & 
           (df.CitypairDestinationCountryCode =='SE'), 'DomesticOrInternational'] = 'Domestic'
    df['Today']=Today
    df['Today']=df.Today.astype("datetime64[s]")
    df['FirstDateOfWeek'] = ((df.Today.dt.to_period('W').apply(lambda r: r.start_time).dt.date)
    - pd.DateOffset(days=1)- pd.DateOffset(weeks=1))+pd.DateOffset(days=1)
#    df['LastDateOfWeek'] = (df.FirstDateOfWeek+ pd.DateOffset(days=7))-pd.DateOffset(days=1)
#    df['IsPreviousWeek']=0
#    df.loc[(df.Date >=df.FirstDateOfWeek) & (df.Date<=df.LastDateOfWeek), 'IsPreviousWeek'] = 1
    df["MonthNameYearEnglish"] = df.Date.dt.strftime('%B')+' '+ df.Date.dt.strftime('%Y').astype(str)
    df["DepartureMonthCode"] = df.Date.dt.strftime('%Y%m')
    df.rename(columns={'Date':'DepartureDate'},inplace=True)
    df.to_sql(name=table,con=SQLITE3,index=False,if_exists='append')


    
''' ..........................................................'''

table='vfTicketCoupons'
sql_query= """

SELECT distinct *
FROM
    pbiFB.vfTicketCoupons vtc
WHERE
vtc.DepartureDate_ID>=%s and
     (vtc.IsIZRClass = 'False') AND
     (vtc.IsCrew = 'False')   
"""%(period_start)

pd.io.sql.execute('DROP TABLE IF EXISTS '+ table, SQLITE3)
for df in pd.read_sql_query(sql_query, SQLServer,chunksize=chunksize):
    df.to_sql(name=table,con=SQLITE3,index=False,if_exists='append')

table='fTicketCoupons'
sql_query= """

SELECT 
       vtc.DepartureDate_ID,
       vtc.BookedDate_ID,
       vcp.CitypairBKEY,
       vcp.CitypairName,
       vcp.CitypairSortOrder,
       vcp.CitypairOriginCountryCode,
       vcp.CitypairDestinationCountryCode,
       ver.ExchangeRate,
       vtc.Amount
FROM
    vfTicketCoupons vtc
    INNER JOIN vdCitypair vcp ON vcp.Citypair_ID = vtc.Citypair_ID
    INNER JOIN vdExchangeRate ver ON ver.CurrencyBKEY = vtc.CurrencyBKEY
"""

pd.io.sql.execute('DROP TABLE IF EXISTS '+ table, SQLITE3)
for df in pd.read_sql_query(sql_query, SQLITE3,chunksize=chunksize):
    df['Date'] =(df.DepartureDate_ID.astype(str).str[0:4]+'-'+
    df.DepartureDate_ID.astype(str).str[4:6]+'-'+
    df.DepartureDate_ID.astype(str).str[6:8]).values.astype("datetime64[s]")
    df['DomesticOrInternational']='International'
    df.loc[(df.CitypairOriginCountryCode == 'SE') & 
           (df.CitypairDestinationCountryCode =='SE'), 'DomesticOrInternational'] = 'Domestic'
    df['Today']=Today
    df['Today']=df.Today.astype("datetime64[s]")
    df['FirstDateOfWeek'] = ((df.Today.dt.to_period('W').apply(lambda r: r.start_time).dt.date)
    - pd.DateOffset(days=1)- pd.DateOffset(weeks=1))+pd.DateOffset(days=1)
    df['LastDateOfWeek'] = (df.FirstDateOfWeek+ pd.DateOffset(days=7))-pd.DateOffset(days=1)
    df['BookedDate'] =(df.BookedDate_ID.astype(str).str[0:4]+'-'+
    df.BookedDate_ID.astype(str).str[4:6]+'-'+
    df.BookedDate_ID.astype(str).str[6:8]).values.astype("datetime64[s]")
    df['IsPreviousWeek']=0
    df.loc[(df.BookedDate >=df.FirstDateOfWeek) & (df.BookedDate<=df.LastDateOfWeek), 'IsPreviousWeek'] = 1
    df["MonthNameYearEnglish"] = df.Date.dt.strftime('%B')+' '+ df.Date.dt.strftime('%Y').astype(str)
    df["DepartureMonthCode"] = df.Date.dt.strftime('%Y%m')
    df['AmountLCY']=df.Amount
    df['AmountLCY'] = pd.to_numeric(df['AmountLCY'], errors='coerce')
    df['ExchangeRate'] = pd.to_numeric(df['ExchangeRate'], errors='coerce')
    df['AmountSEK']=df.AmountLCY*df.ExchangeRate
    df['Bdkpax']=1
    df['AmountBLW']=0
    df.loc[(df.IsPreviousWeek==1) , 'AmountBLW'] = df.AmountSEK
    df['BdkpaxBLW']=0
    df.loc[(df.IsPreviousWeek==1) , 'BdkpaxBLW'] = 1
    df['BookedDateLY'] = df.BookedDate - pd.DateOffset(days=364)
    df['TodayLYx'] = df.Today - pd.DateOffset(days=364)
    df['TodayLY']=df.TodayLYx - pd.DateOffset(weeks=1)
    df['FirstDateOfWeekLY'] = ((df.TodayLY.dt.to_period('W').apply(lambda r: r.start_time).dt.date)
    - pd.DateOffset(days=1)- pd.DateOffset(weeks=0))+pd.DateOffset(days=1)
    df['LastDateOfWeekLY'] = (df.FirstDateOfWeekLY+ pd.DateOffset(days=7))-pd.DateOffset(days=1)
    df['IsPreviousWeekLY']=0
    df.loc[(df.BookedDateLY >=df.FirstDateOfWeekLY) & (df.BookedDateLY<=df.LastDateOfWeekLY), 'IsPreviousWeekLY'] = 1
        #räkna max min DepartureDate baserad på IsPreviousWeekLY.. 
    #df[x]=if IsPreviousWeekLY==1 give me (min ,max DepartureDate_ID )
    #if DepartureDate_IDLY >= then amount
    df.loc[(df.IsPreviousWeekLY==1) , 'BookedDate_IDLY'] = df.BookedDate_ID
    df['AmountBLWLY']=0
    df.loc[(df.BookedDate_IDLY>0) , 'AmountBLWLY'] = 0.1
    df['BdkpaxBLWLY']=0
    df.loc[(df.BookedDate_IDLY>0) , 'BdkpaxBLWLY'] = 0.1
    df.rename(columns={'Date':'DepartureDate'},inplace=True)
    df.to_sql(name=table,con=SQLITE3,index=False,if_exists='append')
    
##
##
#df = pd.read_sql_query(sql_query, SQLITE3)
#df1=pd.DataFrame()
#df1=df
#df1=df1.query('(BookedDate_ID >=20190201 and BookedDate_ID<=20190224)')
#df1['Today']=Today
#df1['Today']=df1.Today.astype("datetime64[s]")
#df1['TodayLYx'] = df1.Today - pd.DateOffset(days=364)
#df1['TodayLY']=df1.TodayLYx - pd.DateOffset(weeks=1)
#df1.drop_duplicates(['BookedDate_ID'],keep= 'last', inplace=True)
#df1 = df1[['DepartureDate','BookedDate','BookedDateLY','TodayLY','IsPreviousWeekLY','IsPreviousWeek','AmountSEK','AmountBLWLY','AmountBLW']]






#
#df['FirstDateOfWeek'] = ((df.Today.dt.to_period('W').apply(lambda r: r.start_time).dt.date)- pd.DateOffset(days=1)- pd.DateOffset(weeks=52))+pd.DateOffset(days=1)
#df['PreviousDateWeek'] = df.Date - pd.DateOffset(days=364)
    
df = pd.DataFrame(index = pd.date_range(period_startDate, period_EndDate, freq='M'))
df['Date']=df.index
df["MonthNameYearEnglish"] = df.Date.dt.strftime('%B')+' '+ df.Date.dt.strftime('%Y').astype(str)
df["DepartureMonthCode"] = df.Date.dt.strftime('%Y%m')

periods=list(df.MonthNameYearEnglish)
periodsCodes=list(df.DepartureMonthCode)

table='hTable'
sql_query='''
WITH
    q1 AS (SELECT DISTINCT
                  ftc.CitypairBKEY
           FROM
               fTicketCoupons ftc where ftc.DepartureDate_ID between %s and %s),
    q2 AS (SELECT
                 cvp.CitypairBKEY,
                 cvp.CitypairSortOrder,
                 cvp.Citypairname,
                 cvp.CitypairOriginCountryCode,
                 cvp.CitypairDestinationCountryCode
           FROM
               vdCitypair cvp)
SELECT
      q1.CitypairBKEY,
      q2.Citypairname,
      q2.CitypairSortOrder,
      q2.CitypairOriginCountryCode,
      q2.CitypairDestinationCountryCode
FROM
    q2
    INNER JOIN q1 ON q1.CitypairBKEY = q2.CitypairBKEY
    
    '''%(period_start_hTable,period_end_hTable)

pd.io.sql.execute('DROP TABLE IF EXISTS '+ table, SQLITE3)
for period,periodsCode in zip(periods,periodsCodes):
    df = pd.read_sql_query(sql_query, SQLITE3)
    df['DomesticOrInternational']='International'
    df.loc[(df.CitypairOriginCountryCode == 'SE') & 
       (df.CitypairDestinationCountryCode =='SE'), 'DomesticOrInternational'] = 'Domestic'
    df['MonthNameYearEnglish']=period
    df["DepartureMonthCode"] = periodsCode  
    df.to_sql(name=table,con=SQLITE3,index=False,if_exists='append')



table='master'
sql_query= """

WITH
    q1 AS (SELECT
      Sum(ftc.AmountSEK) AS BkdRev,
      Sum(ftc.Bdkpax) AS BkdPax,
      Sum(ftc.AmountBLW) AS BkdRevWeek,
      Sum(ftc.BdkpaxBLW) AS BkdPaxWeek,
      Sum(ftc.AmountBLWLY) AS BkdRevWeekLY,
      Sum(ftc.BdkpaxBLWLY) AS BkdPaxWeekLY,
      ht.CitypairBKEY,
      ht.CitypairName,
      ht.CitypairSortOrder,
      ht.DomesticOrInternational,
      ht.MonthNameYearEnglish,
      ht.DepartureMonthCode
FROM
    hTable ht
    LEFT JOIN fTicketCoupons ftc ON (ht.CitypairBKEY = ftc.CitypairBKEY) AND
            (ht.DepartureMonthCode = ftc.DepartureMonthCode)
GROUP BY
        ht.CitypairBKEY,
        ht.CitypairName,
        ht.CitypairSortOrder,
        ht.DomesticOrInternational,
        ht.MonthNameYearEnglish,
        ht.DepartureMonthCode),
    q2 AS (SELECT
                 ff.CitypairBKEY,
                 ff.CitypairName,
                 ff.DepartureMonthCode,
                 ff.MonthNameYearEnglish,
                 Sum(ff.SeatCountTotal) AS Capacity,
                 ff.DomesticOrInternational
           FROM
               fFlights ff
           GROUP BY
                   ff.CitypairBKEY,
                   ff.CitypairName,
                   ff.DepartureMonthCode,
                   ff.MonthNameYearEnglish,
                   ff.DomesticOrInternational),
    q3 AS (SELECT
                 q1.CitypairBKEY,
                 q1.CitypairName,
                 q1.DepartureMonthCode,
                 q1.MonthNameYearEnglish,
                 q1.BkdRev,
                 q2.Capacity,
                 q1.BkdPax,
                 q1.CitypairSortOrder,
                 q1.DomesticOrInternational,
                 q1.BkdRevWeek,
                 q1.BkdPaxWeek,
                 q1.BkdRevWeekLY,
                 q1.BkdPaxWeekLY
           FROM
               q1
               left JOIN q2 ON (q2.CitypairBKEY = q1.CitypairBKEY) AND
                        (q2.DepartureMonthCode = q1.DepartureMonthCode) AND
                        (q2.MonthNameYearEnglish = q1.MonthNameYearEnglish) AND
                        (q2.DomesticOrInternational = q1.DomesticOrInternational))
SELECT
      q3.CitypairBKEY,
      q3.CitypairName,
      q3.DomesticOrInternational,
      q3.CitypairSortOrder,
      q3.DepartureMonthCode,
      q3.MonthNameYearEnglish,
      q3.BkdRev,
      q3.Capacity,
      q3.BkdPax,
      q3.BkdRevWeek,
      q3.BkdPaxWeek,
      q3.BkdRevWeekLY,
      q3.BkdPaxWeekLY
FROM
    q3
"""
pd.io.sql.execute('DROP TABLE IF EXISTS '+ table, SQLITE3)
for df in pd.read_sql_query(sql_query, SQLITE3,chunksize=chunksize):
    df['Loadfact']=df.BkdPax/df.Capacity
    df['BkdRev_Pax']=df.BkdRev/df.BkdPax
    df['BkdRev_PaxWeek']=df.BkdRevWeek/df.BkdPaxWeek
    df['Date'] =(df.DepartureMonthCode.astype(str).str[0:4]+'-'+
    df.DepartureMonthCode.astype(str).str[4:6]+'-01').values.astype("datetime64[s]")
    df['BkdRevLY'] = df.groupby([df['Date'].dt.month,df['Date'].dt.day])['BkdRev'].shift()
    df['CapacityLY'] = df.groupby([df['Date'].dt.month,df['Date'].dt.day])['Capacity'].shift()
    df['BkdPaxLY'] = df.groupby([df['Date'].dt.month,df['Date'].dt.day])['BkdPax'].shift()
    df['LoadfactLY'] = df.groupby([df['Date'].dt.month,df['Date'].dt.day])['Loadfact'].shift()
    df['BkdRev_PaxLY'] = df.groupby([df['Date'].dt.month,df['Date'].dt.day])['BkdRev_Pax'].shift()
    df['BkdRev_PaxWeekLY']=0.1
    df.to_sql(name=table,con=SQLITE3,index=False,if_exists='append')


''' ..................... to excel ............................'''
variables = locals()
table='To_Excel'

for num, period in enumerate(period_to_excel, start=1):
    sql_query='''
WITH
    q1 AS (SELECT
                 hp.CitypairBKEY AS CitypairBKEY,
                 hp.CitypairName,
                 hp.DomesticOrInternational,
                 hp.CitypairSortOrder AS CitypairSortOrder,
                 hp.MonthNameYearEnglish,
                 hp.DepartureMonthCode,
                 Sum(m.BkdRev) AS BkdRev,
                 Sum(m.Capacity) AS Capacity,
                 Sum(m.BkdPax) AS BkdPax,
                 Sum(m.BkdPax) / Sum(m.Capacity) AS Loadfact,
                 Sum(m.BkdRev) / Sum(m.BkdPax) AS BkdRev_Pax,
                 Sum(m.BkdRevWeek) AS BkdRevWeek,
                 Sum(m.BkdPaxWeek) AS BkdPaxWeek,
                 Sum(m.BkdRevWeek) / Sum(m.BkdPaxWeek) AS BkdRev_PaxWeek,
                 sum(m.BkdRevWeekLY) as BkdRevWeekLY,
                 sum(m.BkdPaxWeekLY) as BkdPaxWeekLY,
                 sum(m.BkdRevLY) as BkdRevLY,
                  sum(m.CapacityLY) as CapacityLY,
                  sum(m.BkdPaxLY) as BkdPaxLY,
                  sum(m.LoadfactLY) as LoadfactLY, 
                  sum(m.BkdRev_PaxLY) as BkdRev_PaxLY
           FROM
               hTable hp
               LEFT JOIN master m ON (hp.CitypairBKEY = m.CitypairBKEY) AND
                       (hp.MonthNameYearEnglish = m.MonthNameYearEnglish)
           WHERE
                hp.DomesticOrInternational = 'Domestic'
           GROUP BY
                   hp.CitypairBKEY,
                   hp.CitypairName,
                   hp.DomesticOrInternational,
                   hp.CitypairSortOrder,
                   hp.MonthNameYearEnglish,
                   hp.DepartureMonthCode
                   ),
    q2 AS (SELECT
                 'Domestic' AS CitypairBKEY,
                 'Domestic' as CitypairName,
                 hp.DomesticOrInternational,
                 100000 AS CitypairSortOrder,
                 hp.MonthNameYearEnglish,
                 hp.DepartureMonthCode,
                 Sum(m.BkdRev) AS BkdRev,
                 Sum(m.Capacity) AS Capacity,
                 Sum(m.BkdPax) AS BkdPax,
                 Sum(m.BkdPax) / Sum(m.Capacity) AS Loadfact,
                 Sum(m.BkdRev) / Sum(m.BkdPax) AS BkdRev_Pax,
                 Sum(m.BkdRevWeek) AS BkdRevWeek,
                 Sum(m.BkdPaxWeek) AS BkdPaxWeek,
                 Sum(m.BkdRevWeek) / Sum(m.BkdPaxWeek) AS BkdRev_PaxWeek,
                 sum(m.BkdRevWeekLY) as BkdRevWeekLY,
                 sum(m.BkdPaxWeekLY) as BkdPaxWeekLY,
                 sum(m.BkdRevLY) as BkdRevLY,
                  sum(m.CapacityLY) as CapacityLY,
                  sum(m.BkdPaxLY) as BkdPaxLY,
                  sum(m.LoadfactLY) as LoadfactLY, 
                  sum(m.BkdRev_PaxLY) as BkdRev_PaxLY
           FROM
               hTable hp
               LEFT JOIN master m ON (hp.CitypairBKEY = m.CitypairBKEY) AND
                       (hp.MonthNameYearEnglish = m.MonthNameYearEnglish)
           WHERE
                hp.DomesticOrInternational = 'Domestic'
           GROUP BY
                   hp.DomesticOrInternational,
                   hp.MonthNameYearEnglish,
                   hp.DepartureMonthCode),
    q3 AS (SELECT
                 hp.CitypairBKEY AS CitypairBKEY,
                 hp.CitypairName,
                 hp.DomesticOrInternational,
                 hp.CitypairSortOrder AS CitypairSortOrder,
                 hp.MonthNameYearEnglish,
                 hp.DepartureMonthCode,
                 Sum(m.BkdRev) AS BkdRev,
                 Sum(m.Capacity) AS Capacity,
                 Sum(m.BkdPax) AS BkdPax,
                 Sum(m.BkdPax) / Sum(m.Capacity) AS Loadfact,
                 Sum(m.BkdRev) / Sum(m.BkdPax) AS BkdRev_Pax,
                 Sum(m.BkdRevWeek) AS BkdRevWeek,
                 Sum(m.BkdPaxWeek) AS BkdPaxWeek,
                 Sum(m.BkdRevWeek) / Sum(m.BkdPaxWeek) AS BkdRev_PaxWeek,
                 sum(m.BkdRevWeekLY) as BkdRevWeekLY,
                 sum(m.BkdPaxWeekLY) as BkdPaxWeekLY,
                 sum(m.BkdRevLY) as BkdRevLY,
                  sum(m.CapacityLY) as CapacityLY,
                  sum(m.BkdPaxLY) as BkdPaxLY,
                  sum(m.LoadfactLY) as LoadfactLY, 
                  sum(m.BkdRev_PaxLY) as BkdRev_PaxLY
           FROM
               hTable hp
               LEFT JOIN master m ON (hp.CitypairBKEY = m.CitypairBKEY) AND
                       (hp.MonthNameYearEnglish = m.MonthNameYearEnglish)
           WHERE
                hp.DomesticOrInternational = 'International'
           GROUP BY
                   hp.CitypairBKEY,
                   hp.CitypairName,
                   hp.DomesticOrInternational,
                   hp.CitypairSortOrder,
                   hp.MonthNameYearEnglish),
    q4 AS (SELECT
                 'International' AS CitypairBKEY,
                 'International' as CitypairName,
                 hp.DomesticOrInternational,
                 1000000 AS CitypairSortOrder,
                 hp.MonthNameYearEnglish,
                 hp.DepartureMonthCode,
                 Sum(m.BkdRev) AS BkdRev,
                 Sum(m.Capacity) AS Capacity,
                 Sum(m.BkdPax) AS BkdPax,
                 Sum(m.BkdPax) / Sum(m.Capacity) AS Loadfact,
                 Sum(m.BkdRev) / Sum(m.BkdPax) AS BkdRev_Pax,
                 Sum(m.BkdRevWeek) AS BkdRevWeek,
                 Sum(m.BkdPaxWeek) AS BkdPaxWeek,
                 Sum(m.BkdRevWeek) / Sum(m.BkdPaxWeek) AS BkdRev_PaxWeek,
                 sum(m.BkdRevWeekLY) as BkdRevWeekLY,
                 sum(m.BkdPaxWeekLY) as BkdPaxWeekLY,
                 sum(m.BkdRevLY) as BkdRevLY,
                  sum(m.CapacityLY) as CapacityLY,
                  sum(m.BkdPaxLY) as BkdPaxLY,
                  sum(m.LoadfactLY) as LoadfactLY, 
                  sum(m.BkdRev_PaxLY) as BkdRev_PaxLY
           FROM
               hTable hp
               LEFT JOIN master m ON (hp.CitypairBKEY = m.CitypairBKEY) AND
                       (hp.MonthNameYearEnglish = m.MonthNameYearEnglish)
           WHERE
                hp.DomesticOrInternational = 'International'
           GROUP BY
                   hp.DomesticOrInternational,
                   hp.MonthNameYearEnglish,
                   hp.DepartureMonthCode),
    q5 AS (SELECT
                 'Total' AS CitypairBKEY,
                 'Total' as CitypairName,
                 'Total' AS DomesticOrInternational,
                 10000000 AS CitypairSortOrder,
                 hp.MonthNameYearEnglish,
                 hp.DepartureMonthCode,
                 Sum(m.BkdRev) AS BkdRev,
                 Sum(m.Capacity) AS Capacity,
                 Sum(m.BkdPax) AS BkdPax,
                 Sum(m.BkdPax) / Sum(m.Capacity) AS Loadfact,
                 Sum(m.BkdRev) / Sum(m.BkdPax) AS BkdRev_Pax,
                 Sum(m.BkdRevWeek) AS BkdRevWeek,
                 Sum(m.BkdPaxWeek) AS BkdPaxWeek,
                 Sum(m.BkdRevWeek) / Sum(m.BkdPaxWeek) AS BkdRev_PaxWeek,
                 sum(m.BkdRevWeekLY) as BkdRevWeekLY,
                 sum(m.BkdPaxWeekLY) as BkdPaxWeekLY,
                 sum(m.BkdRevLY) as BkdRevLY,
                  sum(m.CapacityLY) as CapacityLY,
                  sum(m.BkdPaxLY) as BkdPaxLY,
                  sum(m.LoadfactLY) as LoadfactLY, 
                  sum(m.BkdRev_PaxLY) as BkdRev_PaxLY
           FROM
               hTable hp
               LEFT JOIN master m ON (hp.CitypairBKEY = m.CitypairBKEY) AND
                       (hp.MonthNameYearEnglish = m.MonthNameYearEnglish)
           GROUP BY
                   hp.MonthNameYearEnglish,
                   hp.DepartureMonthCode),
    q6 AS (SELECT
                 *
           FROM
               q1
           UNION
           SELECT
                 *
           FROM
               q2
           UNION
           SELECT
                 *
           FROM
               q3
           UNION
           SELECT
                 *
           FROM
               q4
           UNION
           SELECT
                 *
           FROM
               q5)
SELECT
      *
FROM
    q6 where DepartureMonthCode=%s
        
'''%(period)
    df =pd.read_sql_query(sql_query, SQLITE3)
    df['Period']=period
    df['Summa']=df[['BkdRev','Capacity','BkdPax']].sum(axis=1)
    df.loc[-df['CitypairBKEY'].isin(['Domestic','International','Total']), 'Test'] = 1
    df.loc[(df.Summa==0) & (df.Test==1), 'Test1'] = 1
    df=df.query('(Test1 !=1 )') 
    df.drop(['Summa', 'Test','Test1'], axis=1,inplace=True)
    df.to_sql(name='v'+str(num),con=SQLITE3,index=False,if_exists='replace')
    variables["df{0}".format(num)] =df





df4=pd.merge(df1,df2,how="outer",left_on=["CitypairBKEY"],right_on=["CitypairBKEY"])
df4.loc[df4.CitypairSortOrder_x.isnull(), 'CitypairSortOrder_x'] = df4.CitypairSortOrder_y
df4.loc[df4.CitypairName_x.isnull(), 'CitypairName_x'] = df4.CitypairName_y
df4.loc[df4.MonthNameYearEnglish_x.isnull(), 'MonthNameYearEnglish_x'] = df4.MonthNameYearEnglish_y
df4.loc[df4.DepartureMonthCode_x.isnull(), 'DepartureMonthCode_x'] = df4.DepartureMonthCode_x




df=pd.merge(df4,df3,how="outer",left_on=["CitypairBKEY"],right_on=["CitypairBKEY"])
df.loc[df.CitypairSortOrder_x.isnull(), 'CitypairSortOrder_x'] = df.CitypairSortOrder
df.loc[df.CitypairName_x.isnull(), 'CitypairName_x'] = df.CitypairName
df.loc[df.MonthNameYearEnglish_x.isnull(), 'MonthNameYearEnglish_x'] = df.MonthNameYearEnglish
df.loc[df.DepartureMonthCode_x.isnull(), 'DepartureMonthCode_x'] = df.DepartureMonthCode

df.to_sql(name='TheReport',con=SQLITE3,index=False,if_exists='replace')



del df1, df2, df3, df4

sql_query='''
SELECT
      tr.CitypairName_x as CitypairName,
      tr.BkdRev_x as BkdRev,
      tr.Capacity_x as Capacity,
      tr.BkdPax_x as BkdPax,
      tr.Loadfact_x as Loadfact,
      tr.BkdRev_Pax_x as BkdRev_Pax,
      tr.BkdRevWeek_x as BkdRevWeek,
      tr.BkdPaxWeek_x as BkdPaxWeek,
      tr.BkdRev_PaxWeek_x as BkdRev_PaxWeek,
      '' AS x,
      tr.BkdRev_y as BkdRev1,
      tr.Capacity_y as Capacity1,
      tr.BkdPax_y as BkdPax1,
      tr.Loadfact_y as Loadfact1,
      tr.BkdRev_Pax_y as BkdRev_Pax1,
      tr.BkdRevWeek_y as BkdRevWeek1,
      tr.BkdPaxWeek_y as BkdPaxWeek1,
      tr.BkdRev_PaxWeek_y as BkdRev_PaxWeek1,
      '' AS y,
      tr.BkdRev as BkdRev2,
      tr.Capacity as Capacity2,
      tr.BkdPax as BkdPax2,
      tr.Loadfact as Loadfact2,
      tr.BkdRev_Pax as BkdRev_Pax2,
      tr.BkdRevWeek as BkdRevWeek2,
      tr.BkdPaxWeek as BkdPaxWeek2,
      tr.BkdRev_PaxWeek as BkdRev_PaxWeek2
FROM
    TheReport tr
ORDER BY
        tr.CitypairSortOrder_x
   
'''   
df =pd.read_sql_query(sql_query, SQLITE3)
##......................begin stracture for xlsxfile.........................






##......................create xlsxfile.........................................



sheet_name='FBWv'+str(week)

updated=time.strftime("%Y%m%d")
OutFileName='Outfile'
myExcel=path+'\\Outfiles\\'+OutFileName+'_'+updated+'.xlsx'
writer = pd.ExcelWriter(myExcel, engine='xlsxwriter')
df.to_excel(writer, index=False, sheet_name=sheet_name,startrow=4)
workbook = writer.book
worksheet = writer.sheets[sheet_name]
merge_format  = workbook.add_format({'bold': True,'align': 'center'})

percent_fmt = workbook.add_format({'num_format': '0.0%'})

num_cols=['E:E','N:N','W:W']
for num_col in num_cols:
    worksheet.set_column(num_col, None, percent_fmt)
 

format1 = workbook.add_format({'num_format': '#,##0'})
num_cols=['B:B','C:C','B:B' ,'C:C' ,'D:D', 'F:F', 'G:G', 'H:H', 
          'I:I', 'J:J' ,'K:K', 'L:L', 'M:M', 'O:O', 'P:P' ,
          'Q:Q', 'R:R', 'S:S', 'T:T', 'U:U' ,'V:V', 'X:X', 'Y:Y', 'Z:Z', 'AA:AA']
for num_col in num_cols:
    worksheet.set_column(num_col, 18, format1)

worksheet.set_column('A:A', 20)
worksheet.merge_range('A2:AA2', 'Forward Bookings Weekly Report',merge_format )
worksheet.merge_range('B4:I4', str(Period1_excel),merge_format )
worksheet.merge_range('K4:R4', str(Period2_excel),merge_format ) 
worksheet.merge_range('T4:AA4', str(Period3_excel),merge_format )


writer.save()

'''.............. sheet2.................................................'''

sql_query='''
SELECT
      tr.CitypairName_x AS CitypairName,
      tr.BkdRev_x AS BkdRevTY,
      tr.Capacity_x AS CapacityTY,
      tr.BkdPax_x AS BkdPaxTY,
      tr.Loadfact_x AS LoadfactTY,
      tr.BkdRev_Pax_x AS BkdRev_PaxTY,
      tr.BkdRevLY_x AS BkdRevLY,
      tr.CapacityLY_x AS CapacityLY,
      tr.BkdRev_PaxLY_x AS BkdPaxLY,
      tr.LoadfactLY_x AS LoadfactLY,
      tr.BkdRev_PaxLY_x AS BkdRev_PaxLY,
      
      
      tr.BkdRevWeek_x AS BkdRevWeekTY,
      tr.BkdRevWeekLY_x as BkdRevWeekLY,
      tr.BkdPaxWeek_x AS BkdPaxWeekTY,
      tr.BkdPaxWeekLY_x AS BkdPaxWeekLY,
      tr.BkdRev_PaxWeek_x AS BkdRev_PaxWeekTY,
      '' AS BkdRev_PaxWeekLY,
      '' AS x,
      tr.BkdRev_y AS BkdRevTY1,
      tr.Capacity_y AS CapacityTY1,
      tr.BkdPax_y AS BkdPaxTY1,
      tr.Loadfact_y AS LoadfactTY1,
      tr.BkdRev_Pax_y AS BkdRev_PaxTY1,
      tr.BkdRevLY_y AS BkdRevLY1,
      tr.CapacityLY_y AS CapacityLY1,
      tr.BkdRev_PaxLY_y AS BkdPaxLY1,
      tr.LoadfactLY_y AS LoadfactLY1,
      tr.BkdRev_PaxLY_y AS BkdRev_PaxLY1,
      
      
      tr.BkdRevWeek_y AS BkdRevWeekTY1,
      tr.BkdRevWeekLY_y as BkdRevWeekLY1,
      tr.BkdPaxWeek_y AS BkdPaxWeekTY1,
      tr.BkdPaxWeekLY_y AS BkdPaxWeekLY1,
      tr.BkdRev_PaxWeek_y AS BkdRev_PaxWeekTY1,
      '' AS BkdRev_PaxWeekLY1,
      '' AS y,
      tr.BkdRev AS BkdRevTY2,
      tr.Capacity AS CapacityTY2,
      tr.BkdPax AS BkdPaxTY2,
      tr.Loadfact AS LoadfactTY2,
      tr.BkdRev_Pax AS BkdRev_PaxTY2,
      tr.BkdRevLY AS BkdRevLY2,
      tr.CapacityLY AS CapacityLY2,
      tr.BkdRev_PaxLY AS BkdPaxLY2,
      tr.LoadfactLY AS LoadfactLY2,
      tr.BkdRev_PaxLY AS BkdRev_PaxLY2,
      
      
      tr.BkdRevWeek AS BkdRevWeekTY2,
      tr.BkdRevWeekLY as BkdRevWeekLY2,
      tr.BkdPaxWeek AS BkdPaxWeekTY2,
      tr.BkdPaxWeekLY AS BkdPaxWeekLY2,
      tr.BkdRev_PaxWeek AS BkdRev_PaxWeekTY2,
      '' AS BkdRev_PaxWeekLY2
     
FROM
    thereport tr
ORDER BY
        tr.CitypairSortOrder_x
        

'''


df =pd.read_sql_query(sql_query, SQLITE3)



CitypairName=['CitypairName']

lista=' BkdRevTY CapacityTY BkdPaxTY LoadfactTY \
BkdRev_PaxTY BkdRevLY CapacityLY BkdPaxLY LoadfactLY BkdRev_PaxLY\
BkdRevWeekTY BkdRevWeekLY BkdPaxWeekTY BkdPaxWeekLY BkdRev_PaxWeekTY BkdRev_PaxWeekLY'



keepPeriod = [item for item in lista.split(' ') if  item not in  ['']]

keepPeriod1=CitypairName+keepPeriod


new_list=[columns + '1' for columns in keepPeriod ]
keepPeriod2=CitypairName+new_list

new_list=[columns + '2' for columns in keepPeriod ]
keepPeriod3=CitypairName+new_list



df1=df.loc[:, keepPeriod1].copy()
df2=df.loc[:, keepPeriod2].copy()
df3=df.loc[:, keepPeriod3].copy()

with pd.ExcelWriter(myExcel, engine='openpyxl') as writer:
    writer.book = load_workbook(myExcel)
    df1.to_excel(writer, 'FBRv'+str(week),index=False)
    df2.to_excel(writer, 'FBRv'+str(week),index=False,startrow=47)
    df3.to_excel(writer, 'FBRv'+str(week),index=False,startrow=94)

del df1, df2, df3

sql_query='''
SELECT
      m.MonthNameYearEnglish AS Month,
      Sum(m.BkdRev) AS BkdRevTY,
      Sum(m.Capacity) AS CapacityTY,
      Sum(m.BkdPax) AS BkdPaxTY,
      Sum(m.Loadfact) AS LoadfactTY,
      Sum(m.BkdRev_Pax) AS BkdRev_PaxTY,
      Sum(m.BkdRevLY) AS BkdRevLY,
      Sum(m.CapacityLY) AS CapacityLY,
      Sum(m.BkdPaxLY) AS BkdPaxLY,
      Sum(m.LoadfactLY) AS LoadfactLY,
      Sum(m.BkdRev_PaxLY) AS BkdRev_PaxLY,
      Sum(m.BkdRevWeek) AS BkdRevWeekTY,
      Sum(m.BkdRevWeekLY) AS BkdRevWeekLY,
      Sum(m.BkdPaxWeek) AS BkdPaxWeekTY,
      Sum(m.BkdPaxWeekLY) AS BkdPaxWeekLY,
      Sum(m.BkdRev_PaxWeek) AS BkdRev_PaxWeekTY,
      Sum(m.BkdRev_PaxWeekLY) AS BkdRev_PaxWeekLY
FROM
    master m
GROUP BY
        m.MonthNameYearEnglish
ORDER BY
        m.DepartureMonthCode
        

'''


df =pd.read_sql_query(sql_query, SQLITE3)
df.loc[df.BkdRevTY.isnull(), 'test'] = 1
df=df.query('(test !=1 )')
df.drop(['test'], axis=1,inplace=True) 

with pd.ExcelWriter(myExcel, engine='openpyxl') as writer:
    writer.book = load_workbook(myExcel)
    df.to_excel(writer, 'FBMv'+str(week),index=False)


SQLITE3.close()
SQLServer.close()

print('Script successfully completed')
print('Excelfiles are exported to '+myExcel)